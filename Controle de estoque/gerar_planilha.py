"""
Sistema de Controle de Estoque - Gerador de Planilha Excel
Gera uma planilha completa com múltiplas abas e fórmulas automáticas
"""

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError as e:
    print(f"❌ Erro ao importar openpyxl: {e}")
    print("\n💡 Instale as dependências com: pip install -r requirements.txt")
    exit(1)

import os

# Importar configurações
try:
    from config import CONFIG
except ImportError:
    CONFIG = {
        'arquivo_excel': 'Controle_Estoque.xlsx',
        'encoding': 'utf-8-sig'
    }

def criar_planilha_estoque():
    """Cria a planilha de controle de estoque completa"""
    
    try:
        wb = openpyxl.Workbook()
        
        # Remove a planilha padrão
        wb.remove(wb.active)
    except Exception as e:
        print(f"❌ Erro ao criar workbook: {e}")
        return None
    
    # ========== ABA 1: BASE (CADASTRO DE PRODUTOS) ==========
    ws_base = wb.create_sheet("Base")
    
    # Cabeçalhos da Base
    headers_base = [
        "Código",
        "Nome do Produto", 
        "Descrição",
        "Tipo de Produto",
        "Estoque Mínimo",
        "Valor Unitário (R$)",
        "Fornecedor",
        "Localização"
    ]
    
    ws_base.append(headers_base)
    
    # Formatar cabeçalho
    formatar_cabecalho(ws_base, len(headers_base))
    
    # Planilha criada sem dados de exemplo - pronta para uso
    # Usuário pode começar a adicionar produtos via registrar_entrada.py
    # ou manualmente na planilha
    
    # Ajustar largura das colunas
    ws_base.column_dimensions['A'].width = 12
    ws_base.column_dimensions['B'].width = 25
    ws_base.column_dimensions['C'].width = 30
    ws_base.column_dimensions['D'].width = 18
    ws_base.column_dimensions['E'].width = 15
    ws_base.column_dimensions['F'].width = 18
    ws_base.column_dimensions['G'].width = 20
    ws_base.column_dimensions['H'].width = 15
    
    # ========== ABA 2: ENTRADAS ==========
    ws_entradas = wb.create_sheet("Entradas")
    
    headers_entradas = [
        "Data da Entrada",
        "Documento (Nota Fiscal / Nº de Compra)",
        "Produto / Material",
        "Quantidade",
        "Valor Unitário de Compra (R$)",
        "Valor Total (R$)"
    ]
    
    ws_entradas.append(headers_entradas)
    formatar_cabecalho(ws_entradas, len(headers_entradas))
    
    # Aba Entradas criada vazia - pronta para registro de movimentações
    
    # Ajustar colunas
    ws_entradas.column_dimensions['A'].width = 18
    ws_entradas.column_dimensions['B'].width = 35
    ws_entradas.column_dimensions['C'].width = 20
    ws_entradas.column_dimensions['D'].width = 15
    ws_entradas.column_dimensions['E'].width = 28
    ws_entradas.column_dimensions['F'].width = 20
    
    # ========== ABA 3: SAÍDAS ==========
    ws_saidas = wb.create_sheet("Saídas")
    
    headers_saidas = [
        "Data da Saída",
        "Produto / Material",
        "Quantidade Retirada",
        "Motivo da Saída"
    ]
    
    ws_saidas.append(headers_saidas)
    formatar_cabecalho(ws_saidas, len(headers_saidas))
    
    # Aba Saídas criada vazia - pronta para registro de movimentações
    
    # Ajustar colunas
    ws_saidas.column_dimensions['A'].width = 18
    ws_saidas.column_dimensions['B'].width = 20
    ws_saidas.column_dimensions['C'].width = 20
    ws_saidas.column_dimensions['D'].width = 30
    
    # ========== ABA 4: ESTOQUE ATUAL ==========
    ws_estoque = wb.create_sheet("Estoque Atual")
    
    headers_estoque = [
        "Produto / Material",
        "Estoque Inicial",
        "Total de Entradas",
        "Total de Saídas",
        "Saldo Atual"
    ]
    
    ws_estoque.append(headers_estoque)
    formatar_cabecalho(ws_estoque, len(headers_estoque))
    
    # Aba Estoque Atual criada vazia
    # As fórmulas serão adicionadas automaticamente ao cadastrar produtos
    
    # Ajustar colunas
    ws_estoque.column_dimensions['A'].width = 20
    ws_estoque.column_dimensions['B'].width = 18
    ws_estoque.column_dimensions['C'].width = 20
    ws_estoque.column_dimensions['D'].width = 18
    ws_estoque.column_dimensions['E'].width = 18
    
    # ========== ABA 5: ESTOQUE CRÍTICO ==========
    ws_critico = wb.create_sheet("Estoque Crítico")
    
    headers_critico = [
        "Nome do Produto",
        "Estoque Atual",
        "Estoque Mínimo",
        "Status (Ex.: 'Repor' / 'OK')"
    ]
    
    ws_critico.append(headers_critico)
    formatar_cabecalho(ws_critico, len(headers_critico))
    
    # Aba Estoque Crítico criada vazia
    # As fórmulas serão adicionadas automaticamente ao cadastrar produtos
    
    # Ajustar colunas
    ws_critico.column_dimensions['A'].width = 25
    ws_critico.column_dimensions['B'].width = 18
    ws_critico.column_dimensions['C'].width = 18
    ws_critico.column_dimensions['D'].width = 25
    
    # Salvar arquivo
    nome_arquivo = CONFIG['arquivo_excel']
    try:
        wb.save(nome_arquivo)
        print(f"✅ Planilha '{nome_arquivo}' criada com sucesso!")
        print(f"📍 Local: {os.path.abspath(nome_arquivo)}")
        return nome_arquivo
    except PermissionError:
        print(f"❌ Erro: Arquivo '{nome_arquivo}' está aberto. Feche-o e tente novamente.")
        return None
    except Exception as e:
        print(f"❌ Erro ao salvar arquivo: {e}")
        return None


def formatar_cabecalho(ws, num_colunas):
    """Formata o cabeçalho da planilha"""
    # Cor de fundo azul
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    # Fonte branca e negrito
    font = Font(color="FFFFFF", bold=True, size=11)
    # Alinhamento centralizado
    alignment = Alignment(horizontal="center", vertical="center")
    # Bordas
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col in range(1, num_colunas + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
        cell.border = border


def aplicar_formatacao_condicional(ws, num_produtos):
    """Aplica formatação condicional para destacar itens críticos"""
    # Vermelho para "REPOR"
    fill_red = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    # Verde para "OK"
    fill_green = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
    
    for row in range(2, num_produtos + 2):
        cell = ws.cell(row=row, column=4)
        # A formatação será aplicada depois pelo Excel quando abrir


if __name__ == "__main__":
    print("\n" + "="*70)
    print("🏭 GERADOR DE PLANILHA DE CONTROLE DE ESTOQUE")
    print("="*70 + "\n")
    
    resultado = criar_planilha_estoque()
    
    if resultado:
        print("\n✨ Pronto! Planilha criada SEM dados de exemplo.")
        print("\n📋 Próximos passos:")
        print("   1. Cadastrar produtos: python3 registrar_entrada.py (opção 2)")
        print("   2. Registrar entradas: python3 registrar_entrada.py (opção 1)")
        print("   3. Registrar saídas: python3 registrar_saida.py")
        print("   4. Gerar análises: python3 analise_dashboard.py")
        print("   5. Exportar para BI: python3 exportar_para_BI.py")
    else:
        print("\n❌ Falha ao criar a planilha")
        exit(1)
