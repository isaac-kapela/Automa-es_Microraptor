"""
Funções utilitárias compartilhadas - Sistema de Controle de Estoque
"""
import os
from datetime import datetime

try:
    from config import CONFIG
except ImportError:
    CONFIG = {'arquivo_excel': 'Controle_Estoque.xlsx', 'encoding': 'utf-8-sig'}

def limpar_tela():
    """Limpa a tela do terminal"""
    os.system('clear' if os.name != 'nt' else 'cls')

def validar_data(data_str):
    """Valida formato de data DD/MM/YYYY"""
    try:
        datetime.strptime(data_str, '%d/%m/%Y')
        return True
    except ValueError:
        return False

def obter_data(mensagem="Data (DD/MM/YYYY) [Enter = hoje]"):
    """Obtém data validada do usuário"""
    while True:
        data = input(f"\n📅 {mensagem}: ").strip()
        if not data:
            return datetime.now().strftime("%d/%m/%Y")
        if validar_data(data):
            return data
        print("   ❌ Data inválida! Use DD/MM/YYYY")

def obter_numero(mensagem, permitir_decimal=False, minimo=0):
    """Obtém número validado do usuário"""
    while True:
        valor = input(f"\n{mensagem}: ").strip().replace(',', '.')
        try:
            num = float(valor) if permitir_decimal else int(valor)
            if num > minimo:
                return num
            print(f"   ❌ Valor deve ser maior que {minimo}!")
        except ValueError:
            print("   ❌ Número inválido!")

def confirmar(mensagem="Confirmar"):
    """Solicita confirmação do usuário"""
    return input(f"\n✓ {mensagem}? (s/n): ").strip().lower() == 's'

def pausar():
    """Pausa até usuário pressionar Enter"""
    input("\nPressione Enter para continuar...")

def carregar_produtos():
    """Carrega lista de produtos da planilha"""
    import pandas as pd
    arquivo = CONFIG['arquivo_excel']
    
    if not os.path.exists(arquivo):
        print(f"❌ Arquivo '{arquivo}' não encontrado!")
        print("💡 Execute: python3 gerar_planilha.py")
        return None
    
    try:
        return pd.read_excel(arquivo, sheet_name='Base')
    except Exception as e:
        print(f"❌ Erro ao ler produtos: {e}")
        return None

def gerar_proximo_codigo(df_base):
    """Gera o próximo código disponível automaticamente"""
    if df_base.empty:
        return 'P001'
    
    # Extrai números dos códigos existentes (assume formato P001, P002, etc)
    codigos = df_base['Código'].values
    numeros = []
    for cod in codigos:
        try:
            # Remove 'P' e converte para int
            num = int(''.join(filter(str.isdigit, str(cod))))
            numeros.append(num)
        except:
            continue
    
    if not numeros:
        return 'P001'
    
    proximo = max(numeros) + 1
    return f'P{proximo:03d}'

def cadastrar_novo_produto(df_base, codigo=None):
    """Cadastra um novo produto na base"""
    import openpyxl
    
    print("\n" + "="*70)
    print("➕ CADASTRAR NOVO PRODUTO")
    print("="*70)
    
    # Gerar código automaticamente
    if codigo is None:
        codigo = gerar_proximo_codigo(df_base)
    print(f"\n🏷️  Código: {codigo} (gerado automaticamente)")
    
    # Coletar dados
    nome = input("📦 Nome do Produto: ").strip()
    if not nome:
        print("❌ Nome obrigatório!")
        return None
    
    categoria = input("📂 Categoria/Tipo [Enter = Outros]: ").strip() or "Outros"
    descricao = input("📝 Descrição [Enter = pular]: ").strip() or nome
    valor = obter_numero("💰 Valor unitário (R$)", permitir_decimal=True, minimo=-1)
    est_min = obter_numero("📊 Estoque mínimo", minimo=0)
    fornecedor = input("🏭 Fornecedor [Enter = N/D]: ").strip() or "N/D"
    localizacao = input("📍 Localização [Enter = N/D]: ").strip() or "N/D"
    
    # Confirmar
    print("\n" + "-"*70)
    print(f"Código: {codigo} | Nome: {nome}")
    print(f"Categoria: {categoria} | Valor: R$ {valor:.2f}")
    print(f"Estoque Mín: {est_min} | Fornecedor: {fornecedor}")
    print(f"Localização: {localizacao}")
    print("-"*70)
    
    if not confirmar("Confirmar cadastro"):
        print("\n❌ Cancelado!")
        return None
    
    # Salvar na aba Base
    try:
        arquivo = CONFIG['arquivo_excel']
        wb = openpyxl.load_workbook(arquivo)
        ws_base = wb['Base']
        proxima_linha = ws_base.max_row + 1
        
        ws_base.cell(row=proxima_linha, column=1, value=codigo)
        ws_base.cell(row=proxima_linha, column=2, value=nome)
        ws_base.cell(row=proxima_linha, column=3, value=descricao)
        ws_base.cell(row=proxima_linha, column=4, value=categoria)
        ws_base.cell(row=proxima_linha, column=5, value=est_min)
        ws_base.cell(row=proxima_linha, column=6, value=valor)
        ws_base.cell(row=proxima_linha, column=7, value=fornecedor)
        ws_base.cell(row=proxima_linha, column=8, value=localizacao)
        
        # Adicionar fórmulas na aba Estoque Atual
        ws_estoque = wb['Estoque Atual']
        linha_estoque = ws_estoque.max_row + 1
        ws_estoque.cell(row=linha_estoque, column=1, value=codigo)
        ws_estoque.cell(row=linha_estoque, column=2, value=0)  # Estoque inicial
        ws_estoque.cell(row=linha_estoque, column=3, value=f'=SUMIF(Entradas!C:C,A{linha_estoque},Entradas!D:D)')  # Total entradas
        ws_estoque.cell(row=linha_estoque, column=4, value=f'=SUMIF(Saídas!B:B,A{linha_estoque},Saídas!C:C)')  # Total saídas
        ws_estoque.cell(row=linha_estoque, column=5, value=f'=B{linha_estoque}+C{linha_estoque}-D{linha_estoque}')  # Saldo atual
        
        # Adicionar fórmulas na aba Estoque Crítico
        ws_critico = wb['Estoque Crítico']
        linha_critico = ws_critico.max_row + 1
        ws_critico.cell(row=linha_critico, column=1, value=f'=Base!B{proxima_linha}')  # Nome do produto
        ws_critico.cell(row=linha_critico, column=2, value=f'=VLOOKUP(Base!A{proxima_linha},"Estoque Atual"!A:E,5,FALSE)')  # Estoque atual
        ws_critico.cell(row=linha_critico, column=3, value=f'=Base!E{proxima_linha}')  # Estoque mínimo
        ws_critico.cell(row=linha_critico, column=4, value=f'=IF(B{linha_critico}<C{linha_critico},"⚠️ REPOR","✓ OK")')  # Status
        
        wb.save(arquivo)
        print(f"\n✅ Produto {codigo} cadastrado com sucesso!")
        print(f"   Fórmulas adicionadas automaticamente nas abas de estoque.")
        return codigo
    except PermissionError:
        print(f"\n❌ Arquivo está aberto! Feche e tente novamente.")
        return None
    except Exception as e:
        print(f"\n❌ Erro ao cadastrar: {e}")
        return None

def obter_produto(df_base, permitir_novo=True):
    """Obtém e valida código do produto ou cadastra novo"""
    while True:
        print("\n" + "-"*70)
        print("1 - Selecionar produto existente")
        if permitir_novo:
            print("2 - Cadastrar novo produto")
        print("3 - Ver lista completa")
        print("-"*70)
        
        opcao = input("\nOpção: ").strip()
        
        if opcao == '1':
            codigo = input("\n🏷️  Digite o código do produto: ").strip().upper()
            if not codigo:
                print("   ❌ Código obrigatório!")
                continue
            
            if codigo in df_base['Código'].values:
                nome = df_base[df_base['Código'] == codigo]['Nome do Produto'].values[0]
                print(f"   ✓ {nome}")
                return codigo
            
            print(f"   ❌ Produto '{codigo}' não encontrado!")
        
        elif opcao == '2' and permitir_novo:
            # Cadastrar novo produto
            novo_codigo = cadastrar_novo_produto(df_base)
            if novo_codigo:
                # Recarregar base para incluir o novo produto
                df_base_nova = carregar_produtos()
                if df_base_nova is not None:
                    return novo_codigo
        
        elif opcao == '3':
            print("\n" + "="*70)
            for _, row in df_base.iterrows():
                print(f"  {row['Código']:<10} - {row['Nome do Produto']}")
            print("="*70)
        
        else:
            print("   ❌ Opção inválida!")

def listar_produtos_completo(df_base=None, com_estoque=False):
    """Lista produtos de forma formatada"""
    if df_base is None:
        df_base = carregar_produtos()
    if df_base is None:
        return
    
    print("\n" + "="*70)
    print("📦 PRODUTOS CADASTRADOS")
    print("="*70)
    
    if com_estoque:
        import pandas as pd
        try:
            df_estoque = pd.read_excel(CONFIG['arquivo_excel'], sheet_name='Estoque Atual')
            print(f"\n{'Cód':<8} {'Nome':<30} {'Estoque':<12} {'Mín.':<10}")
            print("-"*70)
            for _, row in df_base.iterrows():
                est = df_estoque[df_estoque['Produto / Material'] == row['Código']]
                saldo = est['Saldo Atual'].values[0] if not est.empty else 0
                status = "⚠️" if saldo < row['Estoque Mínimo'] else "✓"
                print(f"{status} {row['Código']:<6} {row['Nome do Produto']:<30} {saldo:<12.0f} {row['Estoque Mínimo']:<10.0f}")
        except:
            com_estoque = False
    
    if not com_estoque:
        print(f"\n{'Código':<10} {'Nome':<30} {'Categoria':<15} {'Valor (R$)':<12}")
        print("-"*70)
        for _, row in df_base.iterrows():
            print(f"{row['Código']:<10} {row['Nome do Produto']:<30} {row['Tipo de Produto']:<15} {row['Valor Unitário (R$)']:<12.2f}")
    
    print("-"*70)
    print(f"Total: {len(df_base)} produtos\n")

def salvar_na_planilha(aba, dados, arquivo=None):
    """Salva dados em uma aba da planilha"""
    import openpyxl
    if arquivo is None:
        arquivo = CONFIG['arquivo_excel']
    
    try:
        wb = openpyxl.load_workbook(arquivo)
        ws = wb[aba]
        proxima_linha = ws.max_row + 1
        
        for col, valor in enumerate(dados, start=1):
            ws.cell(row=proxima_linha, column=col, value=valor)
        
        wb.save(arquivo)
        return True
    except PermissionError:
        print(f"\n❌ Arquivo '{arquivo}' está aberto! Feche e tente novamente.")
        return False
    except Exception as e:
        print(f"\n❌ Erro ao salvar: {e}")
        return False
